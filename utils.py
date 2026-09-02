"""
دوال مساعدة: حساب نتائج التدقيق، توليد رقم مرجعي، والتصدير إلى Excel/PDF.
"""
import io
import random
import string
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# ---------- ألوان موحّدة لتلوين خلايا الحالة/الأولوية بالجداول (بدون أيقونات) ----------
CORRECTIVE_STATUS_COLORS = {
    "open": "#EF4444", "in_progress": "#F59E0B", "pending_review": "#3B82F6",
    "closed": "#22C55E", "rejected": "#6B7280",
}
PRIORITY_COLORS = {"high": "#EF4444", "medium": "#F59E0B", "low": "#22C55E"}
AUDIT_STATUS_COLORS = {
    "scheduled": "#6B7280", "draft": "#F59E0B", "submitted": "#3B82F6",
    "reviewed": "#8B5CF6", "closed": "#22C55E", "cancelled": "#6B7280",
}
BRANCH_STATUS_COLORS = {"active": "#22C55E", "inactive": "#6B7280", "suspended": "#EF4444"}


def generate_reference() -> str:
    """يولّد رقمًا مرجعيًا فريدًا للتدقيق، مثل AUD-20260818-XJ29"""
    date_part = datetime.utcnow().strftime("%Y%m%d")
    rand_part = "".join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"AUD-{date_part}-{rand_part}"


def calculate_audit_score(answers: list[dict], questions_by_id: dict) -> float:
    """
    يحسب نسبة الالتزام المئوية لتدقيق معين.
    القاعدة: نأخذ الأسئلة القابلة للتطبيق فقط (answer != not_applicable)
    ونحسب: مجموع (الوزن للأسئلة المتوافقة) / مجموع (وزن كل الأسئلة القابلة للتطبيق) * 100
    """
    applicable_weight = 0.0
    earned_weight = 0.0
    for ans in answers:
        q = questions_by_id.get(ans["question_id"])
        if not q:
            continue
        if ans.get("answer") == "not_applicable":
            continue
        applicable_weight += q.weight
        if ans.get("answer") == "compliant":
            earned_weight += q.weight

    if applicable_weight == 0:
        return 0.0
    return round((earned_weight / applicable_weight) * 100, 2)


def score_badge(score: float | None) -> str:
    if score is None:
        return "غير محسوبة"
    return f"{score}%"


def style_status_badges(df: pd.DataFrame, column_color_maps: dict):
    """يلوّن خلايا أعمدة معيّنة (حالة، أولوية...) بخلفية ملوّنة حسب قيمتها،
    بدل النص العادي المسطّح — لجعل الجداول أوضح وأسرع قراءة بلمحة، وبدون أي
    أيقونات (اللون وحده كافٍ للتمييز). column_color_maps: {اسم_العمود: {قيمة: لون}}."""
    styler = df.style

    def _make_fn(color_map):
        def _fn(val):
            color = color_map.get(val)
            if not color:
                return ""
            return f"background-color:{color}; color:white; font-weight:600; border-radius:6px;"
        return _fn

    for col, color_map in column_color_maps.items():
        if col in df.columns:
            styler = styler.map(_make_fn(color_map), subset=[col])
    return styler


def paginate_dataframe(df: pd.DataFrame, key_prefix: str, page_size: int = 15, search_label: str = "بحث"):
    """يعرض صندوق بحث نصي (يبحث بكل الأعمدة النصية) + أزرار تنقّل بين الصفحات،
    ويرجّع الجزء المطلوب عرضه من الجدول فقط. يُستخدم لتفادي بطء الجداول الكبيرة
    وتسهيل الوصول لصف معيّن بدل التمرير اليدوي بجدول طويل."""
    if df.empty:
        return df

    query = st.text_input(search_label, key=f"{key_prefix}_search", placeholder=search_label)
    filtered = df
    if query:
        mask = df.apply(lambda row: row.astype(str).str.contains(query, case=False, na=False).any(), axis=1)
        filtered = df[mask]

    total = len(filtered)
    total_pages = max(1, (total - 1) // page_size + 1)
    page_key = f"{key_prefix}_page"
    if page_key not in st.session_state:
        st.session_state[page_key] = 1
    st.session_state[page_key] = min(st.session_state[page_key], total_pages)

    if total_pages > 1:
        pc1, pc2, pc3 = st.columns([1, 2, 1])
        with pc1:
            if st.button("السابق", key=f"{key_prefix}_prev", disabled=st.session_state[page_key] <= 1):
                st.session_state[page_key] -= 1
                st.rerun()
        with pc2:
            st.markdown(
                f"<div style='text-align:center;'>{st.session_state[page_key]} / {total_pages}"
                f" ({total} نتيجة)</div>", unsafe_allow_html=True,
            )
        with pc3:
            if st.button("التالي", key=f"{key_prefix}_next", disabled=st.session_state[page_key] >= total_pages):
                st.session_state[page_key] += 1
                st.rerun()

    start = (st.session_state[page_key] - 1) * page_size
    return filtered.iloc[start:start + page_size]


def export_audits_to_excel(df: pd.DataFrame) -> bytes:
    """يصدر جدول تدقيقات إلى ملف Excel منسّق ويرجعه كـ bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "التدقيقات"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for _, row in df.iterrows():
        ws.append(list(row))

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_audit_report_pdf(audit_info: dict, answers_rows: list[dict]) -> bytes:
    """يصدر تقرير PDF لتدقيق واحد."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"تقرير تدقيق: {audit_info.get('reference','')}", styles["Title"]))
    elements.append(Spacer(1, 12))

    meta = (
        f"الفرع: {audit_info.get('branch_name','')}<br/>"
        f"المدقق: {audit_info.get('auditor_email','')}<br/>"
        f"الحالة: {audit_info.get('status','')}<br/>"
        f"النتيجة: {audit_info.get('score','—')}%<br/>"
    )
    elements.append(Paragraph(meta, styles["Normal"]))
    elements.append(Spacer(1, 16))

    table_data = [["السؤال", "الإجابة", "الوزن", "ملاحظة"]]
    for r in answers_rows:
        table_data.append([r["question"], r["answer"], str(r["weight"]), r.get("note", "")])

    table = Table(table_data, colWidths=[220, 80, 50, 150])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)

    doc.build(elements)
    return buf.getvalue()

